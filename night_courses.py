import asyncio
import aiohttp
import requests
import time
from datetime import datetime
from openpyxl import Workbook

# All lecture types
lectureLessonTypes = [
    'Lecture',
    'Design Lecture',
    'Packaged Lecture',
    'Seminar-Style Module Class',
    'Sectional Teaching',
]

# All tutorial types
tutorialLessonTypes = [
    'Laboratory',
    'Recitation',
    'Tutorial',
    'Tutorial Type 2',
    'Workshop'
]

async def fetch_course(session, moduleCode, academic_year):
    url = f"https://api.nusmods.com/v2/{academic_year}/modules/{moduleCode}.json"
    async with session.get(url) as response:
        try:
            return await response.json()
        except Exception as e:
            print(f"Error processing module {moduleCode}: {e}")
            return None

async def process_modules(modules, academic_year):
    night_courses_by_semester = {1: [], 2: [], 3: [], 4: []}
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_course(session, module['moduleCode'], academic_year) for module in modules]
        results = await asyncio.gather(*tasks)
        for result in results:
            if result:
                for semester in result['semesterData']:
                    all_lectures = [lesson for lesson in semester['timetable'] if lesson['lessonType'] in lectureLessonTypes]
                    filtered_lectures = [lesson for lesson in semester['timetable'] if int(lesson['startTime']) >= 1800 and lesson['lessonType'] in lectureLessonTypes]
                    all_tutorial = [lesson for lesson in semester['timetable'] if lesson['lessonType'] in tutorialLessonTypes]
                    filtered_tutorial = [lesson for lesson in semester['timetable'] if int(lesson['startTime']) >= 1800 and lesson['lessonType'] in tutorialLessonTypes]
                    if (not all_lectures or any(filtered_lectures)) and (not all_tutorial or any(filtered_tutorial)) and (any(all_lectures) or any(all_tutorial)):
                        night_courses_by_semester[semester['semester']].append([result['moduleCode'], result['moduleCredit'], result['title'], result['faculty'], result['department']])
    return night_courses_by_semester

def create_worksheet(workbook, semester, courses):
    worksheet = workbook.create_sheet(title=f"Semester {semester}")
    worksheet['A1'] = 'Code'
    worksheet['B1'] = 'Units'
    worksheet['C1'] = 'Name'
    worksheet['D1'] = 'Faculty'
    worksheet['E1'] = 'Department'

    for idx, course_info in enumerate(courses, start=2):
        module_code, credit_units, course_name, faculty, department = course_info
        worksheet.cell(row=idx, column=1, value=module_code.strip())
        worksheet.cell(row=idx, column=2, value=float(credit_units.strip()))
        worksheet.cell(row=idx, column=3, value=course_name.strip())
        worksheet.cell(row=idx, column=4, value=faculty.strip())
        worksheet.cell(row=idx, column=5, value=department.strip())

def generate_html(courses_by_semester):
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>NUS Night Courses</title>
        <link rel="stylesheet" href="styles.css">
    </head>
    <body>
        <h1>NUS Night Courses</h1>
    """
    for semester, courses in courses_by_semester.items():
        html_content += f"<h2>Semester {semester}</h2>"
        html_content += f"<div class=\"table-container\">"
        html_content += "<table border='1'><tr><th>Code</th><th>Units</th><th>Name</th><th>Faculty</th><th>Department</th></tr>"
        for course_info in courses:
            module_code, credit_units, course_name, faculty, department= course_info
            html_content += f"<tr><td><a href='https://nusmods.com/courses/{module_code}' target='_blank'>{module_code}</a></td><td>{credit_units}</td><td>{course_name}</td><td>{faculty}</td><td>{department}</td></tr>"
        html_content += "</table>"
        html_content += "</div>"
    
    html_content += """
    </body>
    </html>
    """

    return html_content

def write_html_file(html_content):
    with open("index.html", "w") as file:
        file.write(html_content)

async def main():
    academic_year = '2023-2024'
    response = requests.get(f'https://api.nusmods.com/v2/{academic_year}/moduleInfo.json')
    data = response.json()
    courses_by_semester = await process_modules(data, academic_year)

    workbook = Workbook()
    for semester, courses in courses_by_semester.items():
        create_worksheet(workbook, semester, courses)

    today = datetime.now().strftime("%d_%b")
    output_file = f"NUS_Night_Courses_{academic_year}_CAA_{today}.xlsx"
    del workbook['Sheet']
    workbook.save(output_file)

    html_content = generate_html(courses_by_semester)
    write_html_file(html_content)

if __name__ == "__main__":
    start_time = time.time()
    asyncio.run(main())
    print("--- %s seconds ---" % (time.time() - start_time))

